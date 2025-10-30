# ===== IMPORTS NECESSÁRIOS =====
from flask import Flask, request, send_file, render_template, flash, redirect, url_for, session, jsonify
import io
import os
import json
import calendar
from datetime import datetime
from functions.utils import (
    carregar_dados_json,
    salvar_dados_json,
    carregar_usuarios,
    salvar_usuarios,
    carregar_lotes,
    carregar_unidades,
    gerar_datas_do_mes,
    processar_dados_tabulares,
    processar_dados_siisp,
    migrar_dados_existentes,
    calcular_colunas_siisp,
    salvar_mapas_atualizados,
    carregar_mapas,
    obter_unidades_do_lote,
    obter_mapas_do_lote,
    adicionar_usuario,
    buscar_usuario_por_email_ou_usuario,
    validar_dados_unicos,
    atualizar_acesso_usuario
)
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SGMRP - Sistema de Gerenciamento de Mapas de Refeições Penitenciário
Arquivo principal da aplicação Flask

Autor: SEAP
Data: Outubro 2025
"""
import os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DADOS_DIR = os.path.join(BASE_DIR, 'dados')



# Configuração da aplicação Flask
app = Flask(__name__)
app.secret_key = 'sgmrp_seap_2025_secret_key_desenvolvimento'  # Em produção, usar variável de ambiente
app.config['DEBUG'] = True

# ===== ROTA DE EXPORTAÇÃO DE TABELA EXCEL =====
@app.route('/exportar-tabela')
def exportar_tabela():
    # Receber filtros da query string
    lote_id = request.args.get('lote_id', type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    unidades = request.args.get('unidades')
    unidades_list = unidades.split(',') if unidades else []

    # Carregar mapas do lote
    mapas = carregar_dados_json('mapas.json').get('mapas', [])
    lotes = carregar_lotes()
    lote = next((l for l in lotes if l['id'] == lote_id), None)
    precos = lote.get('precos', {}) if lote else {}

    # Filtrar mapas conforme filtros
    def filtro_mapa(m):
        if m['lote_id'] != lote_id:
            return False
        if unidades_list and m.get('nome_unidade') not in unidades_list:
            return False
        if data_inicio and data_fim:
            # Verifica se há datas dentro do intervalo
            datas = m.get('data', [])
            if not datas:
                return False
            # Assume datas no formato DD/MM/YYYY
            def data_br_to_iso(d):
                d = d.split('/')
                return f"{d[2]}-{d[1].zfill(2)}-{d[0].zfill(2)}"
            datas_iso = [data_br_to_iso(d) for d in datas]
            if not any(data_inicio <= d <= data_fim for d in datas_iso):
                return False
        return True
    mapas_filtrados = [m for m in mapas if filtro_mapa(m)]

    # Gerar Excel
    from openpyxl import load_workbook
    # Carregar modelo.xlsx
    modelo_path = os.path.join(DADOS_DIR, 'modelo.xlsx')
    wb = load_workbook(modelo_path)
    # Seleciona a planilha COMPARATIVO
    if 'COMPARATIVO' in wb.sheetnames:
        ws1 = wb['COMPARATIVO']
    else:
        ws1 = wb.active
        ws1.title = 'COMPARATIVO'

    # Copiar conteúdo da planilha RESUMO do modelo.xlsx para a saída
    if 'RESUMO' in wb.sheetnames:
        ws_resumo_saida = wb['RESUMO']
        # Adiciona o texto "Olá, Mundo!" na célula A1 da planilha RESUMO
        ws_resumo_saida['A1'] = "Olá, Mundo!"

    # Preencher preços do lote nas células M6 até T6
    # Ordem: Café Interno, Café Func., Almoço Interno, Almoço Func., Lanche Interno, Lanche Func., Jantar Interno, Jantar Func.
    precos_ordem = [
        ('cafe', 'interno'),
        ('cafe', 'funcionario'),
        ('almoco', 'interno'),
        ('almoco', 'funcionario'),
        ('lanche', 'interno'),
        ('lanche', 'funcionario'),
        ('jantar', 'interno'),
        ('jantar', 'funcionario')
    ]
    col_inicio = 13  # M = 13
    from copy import copy
    for idx, (ref, tipo) in enumerate(precos_ordem):
        col = col_inicio + idx
        valor_preco = precos.get(ref, {}).get(tipo, None)
        cell_preco = ws1.cell(row=6, column=col, value=valor_preco)
        # Copiar formatação da célula original
        cell_modelo = ws1.cell(row=6, column=col)
        cell_preco.font = copy(cell_modelo.font)
        cell_preco.border = copy(cell_modelo.border)
        cell_preco.alignment = copy(cell_modelo.alignment)
        cell_preco.number_format = 'General'
        cell_preco.protection = copy(cell_modelo.protection)

    # Buscar cabeçalho 'LOCAÇÃO' nas primeiras 20 linhas
    header = None
    idx_locacao = None
    header_row = None
    for r in range(1, 21):
        row_values = [cell.value for cell in ws1[r]]
        if row_values and 'LOCAÇÃO' in row_values:
            header = row_values
            idx_locacao = row_values.index('LOCAÇÃO')
            header_row = r
            break

    if header is None:
        # Não encontrou cabeçalho, aborta preenchimento
        return ("Cabeçalho LOCAÇÃO não encontrado no modelo.", 400)
    # Se não houver dados filtrados, retorna mensagem
    tem_dados = False

    # Preencher linhas a partir da linha 12 (após cabeçalho mesclado B9:B11)
    linha = 12
    lote_nome = f"LOTE {lote_id}"
    # Captura o estilo da primeira célula das colunas LOCAÇÃO e UNIDADE
    from copy import copy
    locacao_col = idx_locacao + 1
    primeira_locacao = ws1.cell(row=header_row, column=locacao_col)
    locacao_style = {
        'font': copy(primeira_locacao.font),
        'border': copy(primeira_locacao.border),
        'alignment': copy(primeira_locacao.alignment),
        'number_format': primeira_locacao.number_format,
        'protection': copy(primeira_locacao.protection)
    }

    # Detectar índice da coluna UNIDADE
    idx_unidade = None
    for i, col_name in enumerate(header):
        if col_name and str(col_name).strip().upper() == 'UNIDADE':
            idx_unidade = i
            break
    unidade_style = None
    if idx_unidade is not None:
        unidade_col = idx_unidade + 1
        primeira_unidade = ws1.cell(row=header_row, column=unidade_col)
        unidade_style = {
            'font': copy(primeira_unidade.font),
            'border': copy(primeira_unidade.border),
            'alignment': copy(primeira_unidade.alignment),
            'number_format': primeira_unidade.number_format,
            'protection': copy(primeira_unidade.protection)
        }

    # Função para converter data para ISO (YYYY-MM-DD)
    def data_br_to_iso(d):
        d = d.split('/')
        return f"{d[2]}-{d[1].zfill(2)}-{d[0].zfill(2)}"


    # Preencher LOCAÇÃO, UNIDADE e SIISP (coluna C) a partir da linha 12
    linha = 12
    tem_dados = False
    # Copiar formatação da célula A12
    a12 = ws1.cell(row=12, column=1)
    style_a12 = {
        'font': copy(a12.font),
        'border': copy(a12.border),
        'alignment': copy(a12.alignment),
        'number_format': a12.number_format,
        'protection': copy(a12.protection)
    }

    for mapa in mapas_filtrados:
        unidade_nome = mapa.get('nome_unidade', '')
        lote_nome = f"LOTE {lote_id}"
        n_siisp = mapa.get('n_siisp', [])
        datas = mapa.get('data', [])
        # Aplica filtro de unidade
        if unidades_list and unidade_nome not in unidades_list:
            continue
        for i, valor in enumerate(n_siisp):
            # Aplica filtro de data
            if i < len(datas):
                data_iso = data_br_to_iso(datas[i])
                if data_inicio and data_fim:
                    if not (data_inicio <= data_iso <= data_fim):
                        continue
            # LOCAÇÃO (coluna idx_locacao+1), UNIDADE (coluna idx_unidade+1), SIISP (coluna 3)
            if idx_locacao is not None:
                cell_locacao = ws1.cell(row=linha, column=idx_locacao+1, value=lote_nome)
                cell_locacao.font = style_a12['font']
                cell_locacao.border = style_a12['border']
                cell_locacao.alignment = style_a12['alignment']
                cell_locacao.number_format = style_a12['number_format']
                cell_locacao.protection = style_a12['protection']
            if idx_unidade is not None:
                cell_unidade = ws1.cell(row=linha, column=idx_unidade+1, value=unidade_nome)
                cell_unidade.font = style_a12['font']
                cell_unidade.border = style_a12['border']
                cell_unidade.alignment = style_a12['alignment']
                cell_unidade.number_format = style_a12['number_format']
                cell_unidade.protection = style_a12['protection']
            cell_siisp = ws1.cell(row=linha, column=3, value=valor)
            cell_siisp.font = style_a12['font']
            cell_siisp.border = style_a12['border']
            cell_siisp.alignment = style_a12['alignment']
            cell_siisp.number_format = 'General'
            cell_siisp.protection = style_a12['protection']

            # Preencher coluna D (Data) com formatação de A12, salvando como data
            data_val = datas[i] if i < len(datas) else ''
            cell_data = ws1.cell(row=linha, column=4, value=data_val)
            cell_data.font = style_a12['font']
            cell_data.border = style_a12['border']
            cell_data.alignment = style_a12['alignment']
            cell_data.number_format = 'DD/MM/YYYY'  # Formato de data brasileiro
            cell_data.protection = style_a12['protection']

            # Preencher colunas E-L com dados das refeições, todos como números
            colunas_refeicoes = [
                ('cafe_interno', 5),
                ('cafe_funcionario', 6),
                ('almoco_interno', 7),
                ('almoco_funcionario', 8),
                ('lanche_interno', 9),
                ('lanche_funcionario', 10),
                ('jantar_interno', 11),
                ('jantar_funcionario', 12)
            ]
            for campo, col in colunas_refeicoes:
                valor_refeicao = mapa.get(campo, [])[i] if i < len(mapa.get(campo, [])) else None
                cell_refeicao = ws1.cell(row=linha, column=col, value=valor_refeicao)
                cell_refeicao.font = style_a12['font']
                cell_refeicao.border = style_a12['border']
                cell_refeicao.alignment = style_a12['alignment']
                cell_refeicao.number_format = 'General'
                cell_refeicao.protection = style_a12['protection']

            linha += 1
            tem_dados = True

    if not tem_dados:
        return ("Nenhum dado SIISP encontrado para os filtros selecionados.", 404)

    # Copiar fórmulas de M12:T12 para as linhas de dados preenchidas
    linhas_preenchidas = linha - 12  # linha é incrementada após cada preenchimento
    for offset in range(1, linhas_preenchidas):
        target_row = 12 + offset
        for col in range(13, 21):  # M=13, T=20
            formula_or_value = ws1.cell(row=12, column=col).value
            if ws1.cell(row=12, column=col).data_type == 'f':
                # Se for fórmula, ajustar referência de linha
                import re
                def ajusta_formula(formula, linha_origem, linha_destino):
                    return re.sub(r'(\D)'+str(linha_origem)+r'(\D|$)', lambda m: m.group(1)+str(linha_destino)+m.group(2), formula)
                formula_ajustada = ajusta_formula(formula_or_value, 12, target_row)
                ws1.cell(row=target_row, column=col, value=formula_ajustada)
                ws1.cell(row=target_row, column=col).data_type = 'f'
            else:
                ws1.cell(row=target_row, column=col, value=formula_or_value)
            # Copiar formatação
            cell_modelo = ws1.cell(row=12, column=col)
            cell_dest = ws1.cell(row=target_row, column=col)
            from copy import copy
            cell_dest.font = copy(cell_modelo.font)
            cell_dest.border = copy(cell_modelo.border)
            cell_dest.alignment = copy(cell_modelo.alignment)
            cell_dest.number_format = cell_modelo.number_format
            cell_dest.protection = copy(cell_modelo.protection)

    # Copiar regras de formatação condicional de M12:T12 para as linhas de dados preenchidas
    for col in range(13, 21):  # M=13, T=20
        cell_coord = ws1.cell(row=12, column=col).coordinate
        # Coletar regras relevantes sem modificar durante iteração
        regras_para_copiar = []
        for cf_rule in ws1.conditional_formatting:
            if cell_coord in cf_rule.cells:
                for rule in cf_rule.rules:
                    regras_para_copiar.append(rule)
        # Agora replica as regras para as células de destino
        for rule in regras_para_copiar:
            for target_row in range(13, linha):
                target_coord = ws1.cell(row=target_row, column=col).coordinate
                ws1.conditional_formatting.add(target_coord, rule)

    # Salvar em memória e retornar arquivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Nome do arquivo
    nome_arquivo = f"tabela_lote_{lote_id}"
    if data_inicio and data_fim:
        nome_arquivo += f"_{data_inicio}_a_{data_fim}"
    nome_arquivo += ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )

# Dados simulados temporários (até criarmos os JSONs)

# ===== ROTAS DA APLICAÇÃO =====

@app.route('/')
def index():
    """Página inicial - Landing page"""
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        identificador = request.form.get('email', '').strip()  # Pode ser email ou usuário
        senha = request.form.get('senha', '').strip()
        
        # Validação de campos obrigatórios
        if not identificador:
            flash('Por favor, digite seu usuário ou e-mail!', 'error')
            return render_template('login.html')
        
        if not senha:
            flash('Por favor, digite sua senha!', 'error')
            return render_template('login.html')
        
        # Buscar usuário no arquivo JSON (por email ou nome de usuário)
        usuario = buscar_usuario_por_email_ou_usuario(identificador)
        
        if not usuario:
            flash('Usuário não encontrado! Verifique seu e-mail/usuário ou registre-se.', 'error')
            return render_template('login.html')
        
        # Verificar senha
        if usuario['senha'] != senha:
            flash('Senha incorreta! Verifique sua senha e tente novamente.', 'error')
            return render_template('login.html')
        
        # Verificar se usuário tem acesso liberado
        if not usuario.get('acesso', False):
            flash('Sua conta ainda não foi aprovada pelo administrador. Aguarde a liberação ou entre em contato.', 'warning')
            return render_template('login.html')
        
        # Login bem-sucedido
        session['usuario_id'] = usuario['id']
        session['usuario_nome'] = usuario['nome']
        session['usuario_email'] = usuario['email']
        session['usuario_cargo'] = usuario.get('cargo', '')
        session['usuario_usuario'] = usuario.get('usuario', '')
        session['login_sucesso'] = True  # Flag para mostrar mensagem no dashboard
        
        print(f"✅ Login realizado: {usuario['nome']} ({usuario.get('usuario', usuario['email'])})")
        
        # Redirecionar para dashboard
        return redirect(url_for('dashboard'))
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout do sistema"""
    session.clear()
    flash('Logout realizado com sucesso!', 'info')
    return redirect(url_for('index'))

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    """Página de cadastro de usuários"""
    if request.method == 'POST':
        try:
            # Coletar TODOS os dados do formulário
            dados_usuario = {
                'nome': request.form.get('nome', '').strip(),
                'email': request.form.get('email', '').strip().lower(),
                'cpf': request.form.get('cpf', '').strip(),
                'telefone': request.form.get('telefone', '').strip(),
                'cargo': request.form.get('cargo', '').strip(),
                'unidade': request.form.get('unidade', '').strip(),
                'matricula': request.form.get('matricula', '').strip(),
                'usuario': request.form.get('usuario', '').strip().lower(),
                'senha': request.form.get('senha', '').strip(),
                'justificativa': request.form.get('justificativa', '').strip(),
                'aceitarTermos': request.form.get('aceitarTermos', '')
            }
            
            # Validações básicas obrigatórias
            campos_obrigatorios = ['nome', 'email', 'cpf', 'cargo', 'usuario', 'senha', 'justificativa']
            campos_vazios = [campo for campo in campos_obrigatorios if not dados_usuario[campo]]
            
            if campos_vazios:
                campos_nomes = {
                    'nome': 'Nome completo',
                    'email': 'Email',
                    'cpf': 'CPF',
                    'cargo': 'Cargo/Função',
                    'usuario': 'Nome de usuário',
                    'senha': 'Senha',
                    'justificativa': 'Justificativa de acesso'
                }
                nomes_vazios = [campos_nomes.get(campo, campo) for campo in campos_vazios]
                flash(f'Campos obrigatórios não preenchidos: {", ".join(nomes_vazios)}', 'error')
                return render_template('cadastro.html')
            
            # Validar dados únicos (email, CPF, usuário, matrícula, telefone)
            erros_unicidade = validar_dados_unicos(dados_usuario)
            if erros_unicidade:
                for erro in erros_unicidade:
                    flash(erro, 'error')
                return render_template('cadastro.html')
            
            # Verificar se aceitou os termos
            if dados_usuario['aceitarTermos'] != 'on':
                flash('É necessário aceitar os termos de uso para continuar!', 'error')
                return render_template('cadastro.html')
            
            # Adicionar usuário ao arquivo JSON
            novo_usuario = adicionar_usuario(dados_usuario)
            
            if novo_usuario:
                flash(f'Cadastro realizado com sucesso! Aguarde liberação de acesso.', 'success')
                print(f"✅ Novo usuário cadastrado:")
                print(f"   Nome: {novo_usuario['nome']}")
                print(f"   Email: {novo_usuario['email']}")
                print(f"   Usuário: {novo_usuario['usuario']}")
                print(f"   Cargo: {novo_usuario['cargo']}")
                print(f"   Justificativa: {novo_usuario['justificativa']}")
                return redirect(url_for('login'))
            else:
                flash('Erro interno. Tente novamente mais tarde.', 'error')
                
        except Exception as e:
            print(f"❌ Erro no cadastro: {e}")
            flash('Erro interno no cadastro. Tente novamente.', 'error')
    
    return render_template('cadastro.html')

@app.route('/dashboard')
def dashboard():
    """Dashboard principal - requer login"""
    if 'usuario_id' not in session:
        flash('Acesso negado. Faça login primeiro.', 'warning')
        return redirect(url_for('login'))
    
    # Capturar flag de login_sucesso antes de limpar
    mostrar_sucesso = session.pop('login_sucesso', False)
    
    # Carregar lotes e mapas do arquivo JSON
    lotes = carregar_lotes()
    mapas = carregar_mapas()
    
    # Dados para o dashboard
    context = {
        'usuario_nome': session.get('usuario_nome'),
        'mostrar_login_sucesso': mostrar_sucesso,
        'lotes': lotes,
        'total_lotes': len(lotes),
        'lotes_ativos': len([l for l in lotes if l.get('ativo', False)]),
        'total_unidades': sum(len(l.get('unidades', [])) for l in lotes),
        'mapas_dados': mapas  # Passar dados dos mapas para o frontend
    }
    
    return render_template('dashboard.html', **context)

@app.route('/lotes')
def lotes():
    """Página de listagem de lotes"""
    if 'usuario_id' not in session:
        flash('Acesso negado. Faça login primeiro.', 'warning')
        return redirect(url_for('login'))
    
    # Carregar lotes do arquivo JSON
    lotes = carregar_lotes()
    mapas = carregar_mapas()

    def calcular_conformidade_lote(lote):
        lote_id = lote.get('id')
        mapas_lote = [m for m in mapas if m.get('lote_id') == lote_id]
        precos = lote.get('precos', {})
        valor_total = 0
        valor_desvio = 0
        for mapa in mapas_lote:
            # Soma valores esperados multiplicados pelo preço
            campos_precos = [
                ('cafe_interno', precos.get('cafe', {}).get('interno', 0)),
                ('cafe_funcionario', precos.get('cafe', {}).get('funcionario', 0)),
                ('almoco_interno', precos.get('almoco', {}).get('interno', 0)),
                ('almoco_funcionario', precos.get('almoco', {}).get('funcionario', 0)),
                ('lanche_interno', precos.get('lanche', {}).get('interno', 0)),
                ('lanche_funcionario', precos.get('lanche', {}).get('funcionario', 0)),
                ('jantar_interno', precos.get('jantar', {}).get('interno', 0)),
                ('jantar_funcionario', precos.get('jantar', {}).get('funcionario', 0)),
            ]
            for campo, preco in campos_precos:
                valores = mapa.get(campo, [])
                valor_total += sum(valores) * preco if valores else 0
            # Soma desvios SIISP multiplicados pelo preço (considera todos tipos)
            n_siisp = mapa.get('n_siisp', [])
            # Para desvio, soma todos os n_siisp multiplicados pelo preço médio das refeições
            # (No detalhe, soma por tipo, mas aqui só tem n_siisp total, então usa média dos preços)
            if n_siisp:
                # Se possível, soma por tipo igual ao esperado
                # Aqui, para cada campo, se houver campo_siisp, soma os excedentes multiplicados pelo preço
                for idx, campo in enumerate(['cafe_interno', 'cafe_funcionario', 'almoco_interno', 'almoco_funcionario', 'lanche_interno', 'lanche_funcionario', 'jantar_interno', 'jantar_funcionario']):
                    campo_siisp = f"{campo}_siisp"
                    siisp_vals = mapa.get(campo_siisp, [])
                    preco = 0
                    if 'cafe' in campo:
                        preco = precos.get('cafe', {}).get('interno' if 'interno' in campo else 'funcionario', 0)
                    elif 'almoco' in campo:
                        preco = precos.get('almoco', {}).get('interno' if 'interno' in campo else 'funcionario', 0)
                    elif 'lanche' in campo:
                        preco = precos.get('lanche', {}).get('interno' if 'interno' in campo else 'funcionario', 0)
                    elif 'jantar' in campo:
                        preco = precos.get('jantar', {}).get('interno' if 'interno' in campo else 'funcionario', 0)
                    if siisp_vals:
                        # Só soma excedentes positivos
                        valor_desvio += sum([v for v in siisp_vals if v > 0]) * preco
            # Se não houver campos_siisp, soma n_siisp total multiplicado pelo preço médio
            elif n_siisp:
                precos_lista = [p for _, p in campos_precos if p > 0]
                preco_medio = sum(precos_lista) / len(precos_lista) if precos_lista else 1
                valor_desvio += sum(n_siisp) * preco_medio
        if valor_total > 0:
            conformidade = ((valor_total - valor_desvio) / valor_total) * 100
            return round(max(0, conformidade), 1)
        return None

    # Atualiza cada lote com conformidade calculada
    for lote in lotes:
        # Inicialização de variáveis para todos os cálculos
        lote_id = lote.get('id')
        mapas_lote = [m for m in mapas if m.get('lote_id') == lote_id]
        meses_distintos = set()
        precos = lote.get('precos', {})

        # Cálculo de desvio/mês
        valor_desvio_total = 0
        for mapa in mapas_lote:
            for campo, preco in [
                ('cafe_interno_siisp', precos.get('cafe', {}).get('interno', 0)),
                ('cafe_funcionario_siisp', precos.get('cafe', {}).get('funcionario', 0)),
                ('almoco_interno_siisp', precos.get('almoco', {}).get('interno', 0)),
                ('almoco_funcionario_siisp', precos.get('almoco', {}).get('funcionario', 0)),
                ('lanche_interno_siisp', precos.get('lanche', {}).get('interno', 0)),
                ('lanche_funcionario_siisp', precos.get('lanche', {}).get('funcionario', 0)),
                ('jantar_interno_siisp', precos.get('jantar', {}).get('interno', 0)),
                ('jantar_funcionario_siisp', precos.get('jantar', {}).get('funcionario', 0)),
            ]:
                valores = mapa.get(campo, [])
                valor_desvio_total += sum([v for v in valores if v > 0]) * preco if valores else 0
            # Adiciona mês/ano do mapa
            mes = mapa.get('mes')
            ano = mapa.get('ano')
            if mes and ano:
                meses_distintos.add(f"{mes:02d}/{ano}")
        # Cálculo de meses cadastrados
        lote['meses_cadastrados'] = len(meses_distintos)
        if meses_distintos:
            lote['desvio_mes'] = round(valor_desvio_total / len(meses_distintos), 2)
        else:
            lote['desvio_mes'] = 0

        conf = calcular_conformidade_lote(lote)
        lote['conformidade'] = conf if conf is not None else 'N/A'

        # Cálculo de refeições/mês
        total_refeicoes = 0
        for mapa in mapas_lote:
            for campo in [
                'cafe_interno', 'cafe_funcionario',
                'almoco_interno', 'almoco_funcionario',
                'lanche_interno', 'lanche_funcionario',
                'jantar_interno', 'jantar_funcionario'
            ]:
                valores = mapa.get(campo, [])
                total_refeicoes += sum(valores) if valores else 0
            # Adiciona mês/ano do mapa
            mes = mapa.get('mes')
            ano = mapa.get('ano')
            if mes and ano:
                meses_distintos.add(f"{mes:02d}/{ano}")
        # Calcula refeições/mês
        if meses_distintos:
            lote['refeicoes_mes'] = int(total_refeicoes / len(meses_distintos))
        else:
            lote['refeicoes_mes'] = 0

        # Cálculo de custo/mês
        precos = lote.get('precos', {})
        valor_total = 0
        for mapa in mapas_lote:
            for campo, preco in [
                ('cafe_interno', precos.get('cafe', {}).get('interno', 0)),
                ('cafe_funcionario', precos.get('cafe', {}).get('funcionario', 0)),
                ('almoco_interno', precos.get('almoco', {}).get('interno', 0)),
                ('almoco_funcionario', precos.get('almoco', {}).get('funcionario', 0)),
                ('lanche_interno', precos.get('lanche', {}).get('interno', 0)),
                ('lanche_funcionario', precos.get('lanche', {}).get('funcionario', 0)),
                ('jantar_interno', precos.get('jantar', {}).get('interno', 0)),
                ('jantar_funcionario', precos.get('jantar', {}).get('funcionario', 0)),
            ]:
                valores = mapa.get(campo, [])
                valor_total += sum(valores) * preco if valores else 0
        if meses_distintos:
            lote['custo_mes'] = round(valor_total / len(meses_distintos), 2)
        else:
            lote['custo_mes'] = 0

    empresas = sorted(set(lote.get('empresa', '').strip() for lote in lotes if lote.get('empresa')))
    context = {
        'lotes': lotes,
        'unidades': carregar_unidades(),
        'empresas': empresas
    }
    return render_template('lotes.html', **context)

@app.route('/lote/<int:lote_id>')
def lote_detalhes(lote_id):
    """Página de detalhes de um lote específico"""
    if 'usuario_id' not in session:
        flash('Acesso negado. Faça login primeiro.', 'warning')
        return redirect(url_for('login'))
    
    # Carregar lotes e buscar lote específico
    lotes = carregar_lotes()
    lote = next((l for l in lotes if l['id'] == lote_id), None)
    
    if not lote:
        flash('Lote não encontrado!', 'error')
        return redirect(url_for('lotes'))
    
    # Obter unidades do lote com join dos dados
    unidades_lote = obter_unidades_do_lote(lote_id)
    
    # Obter TODOS os mapas do lote (todos os meses disponíveis)
    mapas_lote = obter_mapas_do_lote(lote_id)
    
    context = {
        'lote': lote,
        'unidades_lote': unidades_lote,
        'mapas_lote': mapas_lote
    }
    
    return render_template('lote-detalhes.html', **context)

# ===== ROTAS ADMINISTRATIVAS =====

@app.route('/admin/usuarios')
def admin_usuarios():
    """Página administrativa para gerenciar usuários"""
    if 'usuario_id' not in session:
        flash('Acesso negado. Faça login primeiro.', 'warning')
        return redirect(url_for('login'))
    
    # Verificar se é admin (usuário ID 1)
    if session.get('usuario_id') != 1:
        flash('Acesso negado. Apenas administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    usuarios = carregar_usuarios()
    
    context = {
        'usuarios': usuarios,
        'total_usuarios': len(usuarios),
        'usuarios_pendentes': len([u for u in usuarios if not u.get('acesso', False)]),
        'usuarios_ativos': len([u for u in usuarios if u.get('acesso', False)])
    }
    
    # Por enquanto, retornar dados JSON (depois criaremos template)
    return jsonify(context)

@app.route('/admin/usuarios/<int:user_id>/aprovar', methods=['POST'])
def aprovar_usuario(user_id):
    """Aprovar acesso de um usuário"""
    if 'usuario_id' not in session or session.get('usuario_id') != 1:
        return jsonify({'error': 'Acesso negado'}), 403
    
    usuario = atualizar_acesso_usuario(user_id, True)
    
    if usuario:
        print(f"✅ Acesso aprovado para: {usuario['nome']} ({usuario['email']})")
        return jsonify({
            'success': True, 
            'message': f'Acesso aprovado para {usuario["nome"]}',
            'usuario': usuario
        })
    
    return jsonify({'error': 'Usuário não encontrado'}), 404

@app.route('/admin/usuarios/<int:user_id>/revogar', methods=['POST'])
def revogar_usuario(user_id):
    """Revogar acesso de um usuário"""
    if 'usuario_id' not in session or session.get('usuario_id') != 1:
        return jsonify({'error': 'Acesso negado'}), 403
    
    usuario = atualizar_acesso_usuario(user_id, False)
    
    if usuario:
        print(f"⚠️ Acesso revogado para: {usuario['nome']} ({usuario['email']})")
        return jsonify({
            'success': True, 
            'message': f'Acesso revogado para {usuario["nome"]}',
            'usuario': usuario
        })
    
    return jsonify({'error': 'Usuário não encontrado'}), 404

# ===== ROTAS DE API (JSON) =====

@app.route('/api/adicionar-dados', methods=['POST'])
def api_adicionar_dados():
    """API para adicionar novos dados de mapas"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Receber dados do formulário
        dados = request.get_json()
        
        # Log para debug
        print(f"📥 Dados recebidos na API:")
        print(f"   Dados completos: {dados}")
        
        lote_id = dados.get('lote_id')
        mes = dados.get('mes')
        ano = dados.get('ano')
        nome_unidade = dados.get('unidade')  # Frontend ainda envia como 'unidade'
        texto = dados.get('texto', '')  # Texto extraído do PDF
        dados_siisp = dados.get('dados_siisp', '')  # Dados SIISP opcionais
        
        # Log específico para o campo texto (com escape para ver caracteres especiais)
        print(f"   Campo texto recebido (raw): {repr(texto)}")
        print(f"   Tamanho do texto: {len(texto)}")
        print(f"   Campo dados_siisp recebido: {repr(dados_siisp)}")
        print(f"   Tamanho dados_siisp: {len(dados_siisp)}")
        if len(texto) > 0:
            print(f"   Primeiros 100 chars texto: {repr(texto[:100])}")
        else:
            print(f"   ⚠️ Texto está vazio ou não foi enviado!")
        
        if len(dados_siisp) > 0:
            print(f"   Primeiros 50 chars SIISP: {repr(dados_siisp[:50])}")
        else:
            print(f"   ℹ️ Dados SIISP não fornecidos (campos SIISP ficarão vazios)")
        
        # Validações básicas
        if not lote_id or not mes or not ano or not nome_unidade:
            return jsonify({'error': 'Lote ID, mês, ano e unidade são obrigatórios'}), 400
        
        # Converter para tipos apropriados
        try:
            lote_id = int(lote_id)
            mes = int(mes)
            ano = int(ano)
        except ValueError:
            return jsonify({'error': 'Lote ID, mês e ano devem ser números'}), 400
        
        # Carregar dados existentes do arquivo mapas.json
        dados_mapas = carregar_dados_json('mapas.json')
        if 'mapas' not in dados_mapas:
            dados_mapas['mapas'] = []
        
        # Verificar se já existe registro para esta unidade, mês, ano e lote
        registro_existente_index = None
        for i, registro in enumerate(dados_mapas['mapas']):
            if (registro.get('nome_unidade') == nome_unidade and 
                registro.get('mes') == mes and 
                registro.get('ano') == ano and
                registro.get('lote_id') == lote_id):
                registro_existente_index = i
                break
        
        # Se existe registro para esta unidade/mês/ano/lote, usar o mesmo ID
        if registro_existente_index is not None:
            # Manter o ID do registro existente
            id_a_usar = dados_mapas['mapas'][registro_existente_index].get('id', 1)
            # Remover o registro antigo
            dados_mapas['mapas'].pop(registro_existente_index)
            print(f"🔄 Substituindo registro existente para {nome_unidade} - {mes}/{ano} (Lote {lote_id})")
        else:
            # Gerar novo ID único (baseado no maior ID existente + 1)
            maior_id = 0
            for registro in dados_mapas['mapas']:
                if 'id' in registro and registro['id'] > maior_id:
                    maior_id = registro['id']
            id_a_usar = maior_id + 1
            print(f"✨ Criando novo registro para {nome_unidade} - {mes}/{ano} (Lote {lote_id})")
        
        # Gerar lista de datas do mês
        datas_do_mes = gerar_datas_do_mes(mes, ano)
        dias_esperados = len(datas_do_mes)  # Usar o tamanho da lista de datas como referência
        
        # Processar dados tabulares do campo texto
        dados_refeicoes = processar_dados_tabulares(texto, dias_esperados)
        
        # Processar dados SIISP opcionais
        dados_siisp_processados = processar_dados_siisp(dados_siisp, dias_esperados)
        
        # Validação dos dados de refeições - VERIFICAR ANTES DE SALVAR
        validacao_refeicoes = dados_refeicoes.get('validacao', {})
        validacao_siisp = dados_siisp_processados.get('validacao', {})
        
        # SE HÁ PROBLEMAS DE VALIDAÇÃO EM REFEIÇÕES, NÃO SALVAR E RETORNAR ERRO
        if not validacao_refeicoes.get('valido', True):
            print(f"❌ Dados rejeitados por problemas de validação nas REFEIÇÕES:")
            print(f"   Registros processados: {validacao_refeicoes.get('registros_processados', 0)}")
            print(f"   Dias esperados: {validacao_refeicoes.get('dias_esperados', 0)}")
            print(f"   Mensagem: {validacao_refeicoes.get('mensagem', '')}")
            
            return jsonify({
                'success': False,
                'error': 'Dados de refeições rejeitados por inconsistência',
                'validacao': {
                    'valido': False,
                    'tipo': 'refeicoes',
                    'registros_processados': validacao_refeicoes.get('registros_processados', 0),
                    'dias_esperados': validacao_refeicoes.get('dias_esperados', 0),
                    'mensagem': validacao_refeicoes.get('mensagem', '')
                }
            }), 400
        
        # SE HÁ PROBLEMAS DE VALIDAÇÃO EM SIISP, NÃO SALVAR E RETORNAR ERRO
        if not validacao_siisp.get('valido', True):
            print(f"❌ Dados rejeitados por problemas de validação nos dados SIISP:")
            print(f"   Mensagem: {validacao_siisp.get('mensagem', '')}")
            
            return jsonify({
                'success': False,
                'error': 'Dados SIISP rejeitados por inconsistência',
                'validacao': {
                    'valido': False,
                    'tipo': 'siisp',
                    'mensagem': validacao_siisp.get('mensagem', '')
                }
            }), 400
        
        # SE DADOS VÁLIDOS, PREPARAR E SALVAR
        novo_registro = {
            'id': id_a_usar,
            'lote_id': lote_id,
            'mes': mes,
            'ano': ano,
            'nome_unidade': nome_unidade,
            'data': datas_do_mes,
            'data_criacao': datetime.now().isoformat(),
            'cafe_interno': dados_refeicoes['cafe_interno'],
            'cafe_funcionario': dados_refeicoes['cafe_funcionario'],
            'almoco_interno': dados_refeicoes['almoco_interno'],
            'almoco_funcionario': dados_refeicoes['almoco_funcionario'],
            'lanche_interno': dados_refeicoes['lanche_interno'],
            'lanche_funcionario': dados_refeicoes['lanche_funcionario'],
            'jantar_interno': dados_refeicoes['jantar_interno'],
            'jantar_funcionario': dados_refeicoes['jantar_funcionario'],
            'n_siisp': dados_siisp_processados['n_siisp']
        }
        
        # CALCULAR COLUNAS SIISP AUTOMATICAMENTE (sempre, mesmo sem dados SIISP)
        n_siisp = dados_siisp_processados['n_siisp']
        
        # Se não há dados SIISP, criar lista de zeros
        if not n_siisp:
            n_siisp = [0] * dias_esperados
            novo_registro['n_siisp'] = n_siisp
            print(f"🔢 Criando lista de zeros para n_siisp: {len(n_siisp)} valores")
        
        print(f"🔢 Calculando colunas SIISP automaticamente...")
        
        # Usar a função para calcular as colunas SIISP
        novo_registro = calcular_colunas_siisp(novo_registro)
        
        # Adicionar novo registro
        dados_mapas['mapas'].append(novo_registro)
        
        # Salvar no arquivo mapas.json
        if salvar_dados_json('mapas.json', dados_mapas):
            print(f"✅ Dados salvos com sucesso em mapas.json:")
            print(f"   Lote ID: {lote_id}")
            print(f"   Mês: {mes}, Ano: {ano} ({dias_esperados} dias)")
            print(f"   Unidade: {nome_unidade}")
            print(f"   📊 Refeições: ✅ {validacao_refeicoes.get('registros_processados', 0)} registros processados")
            print(f"   📊 SIISP: ✅ {validacao_siisp.get('mensagem', 'Dados não fornecidos')}")
            
            return jsonify({
                'success': True,
                'message': 'Dados salvos com sucesso!',
                'registro': novo_registro,
                'validacao': {
                    'valido': True,
                    'refeicoes': {
                        'registros_processados': validacao_refeicoes.get('registros_processados', 0),
                        'dias_esperados': validacao_refeicoes.get('dias_esperados', 0),
                        'mensagem': validacao_refeicoes.get('mensagem', '')
                    },
                    'siisp': {
                        'valores_processados': len(dados_siisp_processados['n_siisp']),
                        'mensagem': validacao_siisp.get('mensagem', '')
                    },
                    'mensagem_geral': 'Dados processados e validados com sucesso'
                }
            })
        else:
            return jsonify({'error': 'Erro ao salvar dados'}), 500
            
    except Exception as e:
        print(f"❌ Erro ao adicionar dados: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/excluir-dados', methods=['DELETE'])
def api_excluir_dados():
    """API para excluir dados de mapas"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Receber dados do formulário
        dados = request.get_json()
        
        # Log para debug
        print(f"🗑️ Dados de exclusão recebidos na API:")
        print(f"   Dados completos: {dados}")
        
        lote_id = dados.get('lote_id')
        mes = dados.get('mes')
        ano = dados.get('ano')
        nome_unidade = dados.get('unidade')
        
        # Validações básicas
        if not lote_id or not mes or not ano or not nome_unidade:
            return jsonify({'error': 'Lote ID, mês, ano e unidade são obrigatórios'}), 400
        
        # Converter para tipos apropriados
        try:
            lote_id = int(lote_id)
            mes = int(mes)
            ano = int(ano)
        except ValueError:
            return jsonify({'error': 'Lote ID, mês e ano devem ser números'}), 400
        
        # Carregar dados existentes do arquivo mapas.json
        dados_mapas = carregar_dados_json('mapas.json')
        if 'mapas' not in dados_mapas:
            dados_mapas['mapas'] = []
        
        # Procurar pelo registro específico para excluir
        registro_encontrado = None
        registro_index = None
        
        for i, registro in enumerate(dados_mapas['mapas']):
            if (registro.get('nome_unidade') == nome_unidade and 
                registro.get('mes') == mes and 
                registro.get('ano') == ano and
                registro.get('lote_id') == lote_id):
                registro_encontrado = registro
                registro_index = i
                break
        
        # Verificar se o registro foi encontrado
        if registro_encontrado is None:
            print(f"❌ Registro não encontrado para {nome_unidade} - {mes}/{ano} (Lote {lote_id})")
            return jsonify({
                'success': False,
                'error': f'Registro não encontrado para {nome_unidade} em {mes}/{ano}'
            }), 404
        
        # Remover o registro da lista
        dados_mapas['mapas'].pop(registro_index)
        
        # Salvar dados atualizados
        sucesso_salvamento = salvar_dados_json('mapas.json', dados_mapas)
        
        if sucesso_salvamento:
            print(f"✅ Registro excluído com sucesso:")
            print(f"   Lote ID: {lote_id}")
            print(f"   Período: {mes}/{ano}")
            print(f"   Unidade: {nome_unidade}")
            print(f"   Total de registros restantes: {len(dados_mapas['mapas'])}")
            
            return jsonify({
                'success': True,
                'message': 'Registro excluído com sucesso!',
                'registro_excluido': {
                    'id': registro_encontrado.get('id'),
                    'lote_id': lote_id,
                    'mes': mes,
                    'ano': ano,
                    'nome_unidade': nome_unidade,
                    'data_criacao': registro_encontrado.get('data_criacao')
                }
            })
        else:
            return jsonify({'error': 'Erro ao salvar alterações no arquivo'}), 500
            
    except Exception as e:
        print(f"❌ Erro ao excluir dados: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/entrada-manual', methods=['POST'])
def api_entrada_manual():
    """API para entrada manual de dados de mapas"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Receber dados do formulário
        dados = request.get_json()
        
        # Log para debug
        print(f"✏️ Dados de entrada manual recebidos na API:")
        print(f"   Dados completos: {dados}")
        
        lote_id = dados.get('lote_id')
        mes = dados.get('mes')
        ano = dados.get('ano')
        nome_unidade = dados.get('unidade')
        dados_tabela = dados.get('dados_tabela', [])  # Array com dados da tabela
        
        # Validações básicas
        if not lote_id or not mes or not ano or not nome_unidade:
            return jsonify({'error': 'Lote ID, mês, ano e unidade são obrigatórios'}), 400
        
        if not dados_tabela:
            return jsonify({'error': 'Dados da tabela são obrigatórios'}), 400
        
        # Converter para tipos apropriados  
        try:
            lote_id = int(lote_id)
            mes = int(mes)
            ano = int(ano)
        except ValueError:
            return jsonify({'error': 'Lote ID, mês e ano devem ser números'}), 400
        
        # Carregar dados existentes do arquivo mapas.json
        dados_mapas = carregar_dados_json('mapas.json')
        if 'mapas' not in dados_mapas:
            dados_mapas['mapas'] = []
        
        # Verificar se já existe registro para esta unidade, mês, ano e lote
        registro_existente_index = None
        for i, registro in enumerate(dados_mapas['mapas']):
            if (registro.get('nome_unidade') == nome_unidade and 
                registro.get('mes') == mes and 
                registro.get('ano') == ano and
                registro.get('lote_id') == lote_id):
                registro_existente_index = i
                break
        
        # Se existe registro para esta unidade/mês/ano/lote, usar o mesmo ID
        if registro_existente_index is not None:
            # Manter o ID do registro existente
            id_a_usar = dados_mapas['mapas'][registro_existente_index].get('id', 1)
            # Remover o registro antigo
            dados_mapas['mapas'].pop(registro_existente_index)
            print(f"🔄 Substituindo registro existente para {nome_unidade} - {mes}/{ano} (Lote {lote_id})")
        else:
            # Gerar novo ID único (baseado no maior ID existente + 1)
            maior_id = 0
            for registro in dados_mapas['mapas']:
                if 'id' in registro and registro['id'] > maior_id:
                    maior_id = registro['id']
            id_a_usar = maior_id + 1
            print(f"✨ Criando novo registro para {nome_unidade} - {mes}/{ano} (Lote {lote_id})")
        
        # Gerar lista de datas do mês automaticamente (formato DD/MM/YYYY)
        datas_do_mes = gerar_datas_do_mes(mes, ano)
        dias_esperados = len(datas_do_mes)
        
        # Validar se o número de registros da tabela bate com os dias do mês
        if len(dados_tabela) != dias_esperados:
            return jsonify({
                'error': f'Número de registros ({len(dados_tabela)}) não confere com os dias do mês ({dias_esperados})'
            }), 400
        
        # Processar dados da tabela manual
        # Separar os dados por tipo de refeição
        cafe_interno = []
        cafe_funcionario = []
        almoco_interno = []
        almoco_funcionario = []
        lanche_interno = []
        lanche_funcionario = []
        jantar_interno = []
        jantar_funcionario = []
        
        for dia_dados in dados_tabela:
            # Converter valores para inteiros (0 se vazio ou inválido)
            try:
                cafe_interno.append(int(dia_dados.get('cafe_interno', 0) or 0))
                cafe_funcionario.append(int(dia_dados.get('cafe_funcionario', 0) or 0))
                almoco_interno.append(int(dia_dados.get('almoco_interno', 0) or 0))
                almoco_funcionario.append(int(dia_dados.get('almoco_funcionario', 0) or 0))
                lanche_interno.append(int(dia_dados.get('lanche_interno', 0) or 0))
                lanche_funcionario.append(int(dia_dados.get('lanche_funcionario', 0) or 0))
                jantar_interno.append(int(dia_dados.get('jantar_interno', 0) or 0))
                jantar_funcionario.append(int(dia_dados.get('jantar_funcionario', 0) or 0))
            except (ValueError, TypeError):
                return jsonify({
                    'error': f'Valores inválidos encontrados no dia {dia_dados.get("dia", "?")}'
                }), 400
        
        # Criar novo registro
        novo_registro = {
            'id': id_a_usar,
            'lote_id': lote_id,
            'mes': mes,
            'ano': ano,
            'nome_unidade': nome_unidade,
            'data': datas_do_mes,  # Backend gera automaticamente as datas
            'data_criacao': datetime.now().isoformat(),
            'cafe_interno': cafe_interno,
            'cafe_funcionario': cafe_funcionario,
            'almoco_interno': almoco_interno,
            'almoco_funcionario': almoco_funcionario,
            'lanche_interno': lanche_interno,
            'lanche_funcionario': lanche_funcionario,
            'jantar_interno': jantar_interno,
            'jantar_funcionario': jantar_funcionario,
            'n_siisp': [0] * dias_esperados  # Lista de zeros do tamanho do mês
        }
        
        # CALCULAR COLUNAS SIISP AUTOMATICAMENTE (usando zeros)
        print(f"🔢 Calculando colunas SIISP automaticamente com lista de zeros...")
        novo_registro = calcular_colunas_siisp(novo_registro)
        
        # Adicionar novo registro
        dados_mapas['mapas'].append(novo_registro)
        
        # Salvar no arquivo mapas.json
        if salvar_dados_json('mapas.json', dados_mapas):
            print(f"✅ Entrada manual salva com sucesso em mapas.json:")
            print(f"   Lote ID: {lote_id}")
            print(f"   Mês: {mes}, Ano: {ano} ({dias_esperados} dias)")
            print(f"   Unidade: {nome_unidade}")
            print(f"   📊 Registros processados: {len(dados_tabela)}")
            print(f"   📊 Dados SIISP: Lista vazia (como especificado)")
            
            # Calcular totais para log
            total_refeicoes = sum(cafe_interno + cafe_funcionario + almoco_interno + almoco_funcionario + 
                                lanche_interno + lanche_funcionario + jantar_interno + jantar_funcionario)
            print(f"   📊 Total de refeições: {total_refeicoes}")
            
            return jsonify({
                'success': True,
                'message': 'Dados da entrada manual salvos com sucesso!',
                'registro': novo_registro,
                'estatisticas': {
                    'total_dias': dias_esperados,
                    'total_refeicoes': total_refeicoes,
                    'registros_processados': len(dados_tabela)
                }
            })
        else:
            return jsonify({'error': 'Erro ao salvar dados'}), 500
            
    except Exception as e:
        print(f"❌ Erro na entrada manual: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/adicionar-siisp', methods=['POST'])
def api_adicionar_siisp():
    """API para adicionar números SIISP a registros existentes"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Receber dados do formulário
        dados = request.get_json()
        
        # Log para debug
        print(f"📊 Dados SIISP recebidos na API:")
        print(f"   Dados completos: {dados}")
        
        lote_id = dados.get('lote_id')
        mes = dados.get('mes')
        ano = dados.get('ano')
        nome_unidade = dados.get('unidade')
        numeros_siisp_texto = dados.get('numeros_siisp', '')  # Texto com números SIISP
        
        # Validações básicas
        if not lote_id or not mes or not ano or not nome_unidade:
            return jsonify({'error': 'Lote ID, mês, ano e unidade são obrigatórios'}), 400
        
        if not numeros_siisp_texto or numeros_siisp_texto.strip() == '':
            return jsonify({'error': 'Números SIISP são obrigatórios'}), 400
        
        # Converter para tipos apropriados  
        try:
            lote_id = int(lote_id)
            mes = int(mes)
            ano = int(ano)
        except ValueError:
            return jsonify({'error': 'Lote ID, mês e ano devem ser números'}), 400
        
        # Processar números SIISP (um por linha)
        linhas = numeros_siisp_texto.strip().split('\n')
        numeros_siisp = []
        
        for i, linha in enumerate(linhas, 1):
            linha_limpa = linha.strip()
            if linha_limpa:  # Ignorar linhas vazias
                try:
                    numero = int(linha_limpa)
                    numeros_siisp.append(numero)
                except ValueError:
                    return jsonify({
                        'error': f'Número SIISP inválido na linha {i}: "{linha_limpa}". Deve ser um número inteiro.'
                    }), 400
        
        # Validar quantidade de números vs dias do mês
        dias_esperados = calendar.monthrange(ano, mes)[1]
        if len(numeros_siisp) != dias_esperados:
            return jsonify({
                'error': f'Quantidade de números SIISP ({len(numeros_siisp)}) não confere com os dias do mês {mes}/{ano} ({dias_esperados} dias).'
            }), 400
        
        # Carregar dados existentes do arquivo mapas.json
        dados_mapas = carregar_dados_json('mapas.json')
        if 'mapas' not in dados_mapas:
            dados_mapas['mapas'] = []
        
        # Procurar pelo registro específico para adicionar SIISP
        registro_encontrado = None
        registro_index = None
        
        for i, registro in enumerate(dados_mapas['mapas']):
            if (registro.get('nome_unidade') == nome_unidade and 
                registro.get('mes') == mes and 
                registro.get('ano') == ano and
                registro.get('lote_id') == lote_id):
                registro_encontrado = registro
                registro_index = i
                break
        
        # Verificar se o registro foi encontrado
        if registro_encontrado is None:
            return jsonify({
                'success': False,
                'error': f'Não foram encontrados dados de refeições para {nome_unidade} em {mes}/{ano}. É necessário ter dados de refeições antes de adicionar números SIISP.'
            }), 404
        
        print(f"📊 Registro encontrado para {nome_unidade} - {mes}/{ano} (Lote {lote_id})")
        
        # Atualizar o registro com os números SIISP
        registro_encontrado['n_siisp'] = numeros_siisp
        
        # Calcular automaticamente as colunas SIISP (diferenças)
        print(f"🔢 Calculando colunas SIISP automaticamente...")
        
        campos_refeicoes = [
            'cafe_interno', 'cafe_funcionario',
            'almoco_interno', 'almoco_funcionario', 
            'lanche_interno', 'lanche_funcionario',
            'jantar_interno', 'jantar_funcionario'
        ]
        
        # Para cada tipo de refeição, calcular diferença: valor_refeicao - n_siisp
        for campo in campos_refeicoes:
            campo_siisp = f"{campo}_siisp"
            valores_refeicoes = registro_encontrado.get(campo, [])
            
            if valores_refeicoes and len(valores_refeicoes) == len(numeros_siisp):
                # Calcular diferenças para cada dia
                diferencas = []
                for j in range(len(numeros_siisp)):
                    diferenca = valores_refeicoes[j] - numeros_siisp[j]
                    diferencas.append(diferenca)
                
                registro_encontrado[campo_siisp] = diferencas
                print(f"   ✅ {campo_siisp}: calculado {len(diferencas)} valores")
            else:
                # Se não há dados compatíveis, manter vazio
                registro_encontrado[campo_siisp] = []
                print(f"   ⚠️ {campo_siisp}: dados incompatíveis, mantido vazio")
        
        # Atualizar timestamp de modificação
        registro_encontrado['data_atualizacao_siisp'] = datetime.now().isoformat()
        
        # Salvar dados atualizados
        if salvar_dados_json('mapas.json', dados_mapas):
            print(f"✅ Números SIISP adicionados com sucesso:")
            print(f"   Lote ID: {lote_id}")
            print(f"   Período: {mes}/{ano} ({dias_esperados} dias)")
            print(f"   Unidade: {nome_unidade}")
            print(f"   📊 Números SIISP: {len(numeros_siisp)} valores")
            print(f"   📊 Colunas calculadas: 8 campos de diferenças")
            
            return jsonify({
                'success': True,
                'message': f'Números SIISP adicionados com sucesso para {nome_unidade} em {mes}/{ano}!',
                'registro': registro_encontrado,
                'estatisticas': {
                    'total_dias': dias_esperados,
                    'numeros_adicionados': len(numeros_siisp),
                    'colunas_calculadas': len(campos_refeicoes)
                }
            })
        else:
            return jsonify({'error': 'Erro ao salvar alterações no arquivo'}), 500
            
    except Exception as e:
        print(f"❌ Erro ao adicionar números SIISP: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/lotes')
def api_lotes():
    """API para listar lotes (JSON)"""
    lotes = carregar_lotes()
    return jsonify(lotes)

@app.route('/api/unidades')
def api_unidades():
    """API para listar unidades (JSON)"""
    return jsonify(DADOS_SIMULADOS['unidades'])

@app.route('/api/usuarios')
def api_usuarios():
    """API para listar usuários (apenas para admins)"""
    if 'usuario_id' not in session or session.get('usuario_id') != 1:
        return jsonify({'error': 'Acesso negado'}), 403
    
    usuarios = carregar_usuarios()
    # Remover senhas da resposta por segurança
    usuarios_safe = []
    for u in usuarios:
        usuario_safe = u.copy()
        usuario_safe.pop('senha', None)
        usuarios_safe.append(usuario_safe)
    
    return jsonify(usuarios_safe)

@app.route('/api/validar-campo', methods=['POST'])
def validar_campo_unico():
    """API para validar se um campo específico já existe (validação em tempo real)"""
    try:
        campo = request.json.get('campo')
        valor = request.json.get('valor', '').strip()
        
        if not campo or not valor:
            return jsonify({'valido': True})
        
        usuarios = carregar_usuarios()
        
        # Validações específicas por campo
        existe = False
        mensagem = ''
        
        if campo == 'email':
            valor = valor.lower()
            existe = any(u.get('email', '').lower() == valor for u in usuarios)
            mensagem = 'Este email já está em uso!'
        elif campo == 'cpf':
            existe = any(u.get('cpf', '') == valor for u in usuarios)
            mensagem = 'Este CPF já está cadastrado!'
        elif campo == 'usuario':
            valor = valor.lower()
            existe = any(u.get('usuario', '').lower() == valor for u in usuarios)
            mensagem = 'Este nome de usuário já existe!'
        elif campo == 'matricula':
            if valor:  # Matrícula é opcional
                existe = any(u.get('matricula', '') == valor for u in usuarios)
                mensagem = 'Esta matrícula já está em uso!'
        elif campo == 'telefone':
            if valor:  # Telefone é opcional
                existe = any(u.get('telefone', '') == valor for u in usuarios)
                mensagem = 'Este telefone já está cadastrado!'
        
        return jsonify({
            'valido': not existe,
            'mensagem': mensagem if existe else 'Disponível!'
        })
        
    except Exception as e:
        print(f"Erro na validação: {e}")
        return jsonify({'valido': True}), 500

# ===== FILTROS PERSONALIZADOS PARA TEMPLATES =====

@app.template_filter('data_br')
def filtro_data_br(data_str):
    """Converte data ISO para formato brasileiro"""
    try:
        data = datetime.strptime(data_str, '%Y-%m-%d')
        return data.strftime('%d/%m/%Y')
    except:
        return data_str

@app.template_filter('status_badge')
def filtro_status_badge(status):
    """Retorna classe CSS para badge de status"""
    badges = {
        'concluido': 'success',
        'em_andamento': 'primary', 
        'planejado': 'warning',
        'cancelado': 'danger'
    }
    return badges.get(status, 'secondary')

# ===== CONTEXTO GLOBAL PARA TEMPLATES =====

@app.context_processor
def contexto_global():
    """Variáveis disponíveis em todos os templates"""
    return {
        'app_nome': 'SGMRP',
        'app_versao': '1.0.0',
        'ano_atual': datetime.now().year,
        'usuario_logado': 'usuario_id' in session,
        'usuario_nome': session.get('usuario_nome', ''),
        'usuario_perfil': session.get('usuario_perfil', '')
    }

# ===== TRATAMENTO DE ERROS =====

@app.errorhandler(404)
def pagina_nao_encontrada(error):
    """Página de erro 404"""
    return render_template('index.html'), 404

@app.errorhandler(500)
def erro_interno(error):
    """Página de erro 500"""
    return render_template('index.html'), 500

# ===== INICIALIZAÇÃO DA APLICAÇÃO =====

if __name__ == '__main__':
    # Definir BASE_DIR e DADOS_DIR antes dos prints
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DADOS_DIR = os.path.join(BASE_DIR, 'dados')
    print("🚀 Iniciando SGMRP - Sistema de Gerenciamento de Mapas de Refeições Penitenciário")
    print(f"📁 Diretório base: {BASE_DIR}")
    print(f"📄 Templates: {os.path.join(BASE_DIR, 'templates')}")
    print(f"🎨 Arquivos estáticos: {os.path.join(BASE_DIR, 'static')}")
    print(f"💾 Dados JSON: {DADOS_DIR}")
    
    # Executar migração de dados se necessário
    print("🔄 Verificando migração de dados...")
    migrar_dados_existentes()
    
    print("🔗 Acesse: http://localhost:5000")
    print("👤 Admin: admin@seap.gov.br (ou 'admin') | Senha: admin123")
    print("📋 Usuários: /admin/usuarios (apenas admin)")
    print("📝 Cadastros salvos em: dados/usuarios.json")
    print("-" * 60)
    
    # Executar aplicação
    app.run(
        host='0.0.0.0',      # Aceita conexões de qualquer IP
        port=5000,           # Porta padrão do Flask
        debug=True,          # Modo debug ativo
        use_reloader=True    # Reinicialização automática ao modificar arquivos
    )